from flask import Flask, render_template, request, jsonify, send_file
import json
import os
from datetime import datetime
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

# Config file location (stored in user's home directory, always accessible)
CONFIG_FILE = os.path.expanduser("~/.expense_tracker_config.json")

# Default categories
DEFAULT_CATEGORIES = [
    "Food & Dining",
    "Transportation", 
    "Shopping",
    "Entertainment",
    "Bills & Utilities",
    "Healthcare",
    "Education",
    "Housing",
    "Insurance",
    "Investment",
    "Other"
]

def load_config():
    """Load configuration from file"""
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    else:
        # Default configuration
        default_config = {
            "data_folder": os.path.expanduser("~/Documents/ExpenseTracker")
        }
        save_config(default_config)
        return default_config

def save_config(config):
    """Save configuration to file"""
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)

def get_data_paths():
    """Get all data file paths based on current config"""
    config = load_config()
    data_folder = os.path.expanduser(config['data_folder'])
    
    # Create data folder if it doesn't exist
    os.makedirs(data_folder, exist_ok=True)
    
    return {
        'folder': data_folder,
        'transactions': os.path.join(data_folder, "transactions.json"),
        'budgets': os.path.join(data_folder, "budgets.json"),
        'categories': os.path.join(data_folder, "categories.json")
    }

def load_data():
    """Load transactions from JSON file"""
    paths = get_data_paths()
    if os.path.exists(paths['transactions']):
        with open(paths['transactions'], 'r') as f:
            return json.load(f)
    return []

def save_data(transactions):
    """Save transactions to JSON file"""
    paths = get_data_paths()
    with open(paths['transactions'], 'w') as f:
        json.dump(transactions, f, indent=2)

def load_budgets():
    """Load budgets from JSON file"""
    paths = get_data_paths()
    if os.path.exists(paths['budgets']):
        with open(paths['budgets'], 'r') as f:
            return json.load(f)
    return {}

def save_budgets(budgets):
    """Save budgets to JSON file"""
    paths = get_data_paths()
    with open(paths['budgets'], 'w') as f:
        json.dump(budgets, f, indent=2)

def load_categories():
    """Load custom categories from JSON file"""
    paths = get_data_paths()
    if os.path.exists(paths['categories']):
        with open(paths['categories'], 'r') as f:
            return json.load(f)
    return DEFAULT_CATEGORIES.copy()

def save_categories(categories):
    """Save categories to JSON file"""
    paths = get_data_paths()
    with open(paths['categories'], 'w') as f:
        json.dump(categories, f, indent=2)

@app.route('/')
def index():
    paths = get_data_paths()
    return render_template('index.html', data_folder=paths['folder'])

@app.route('/api/config', methods=['GET'])
def get_config():
    """Get current configuration"""
    config = load_config()
    return jsonify(config)

@app.route('/api/config', methods=['POST'])
def update_config():
    """Update configuration"""
    data = request.json
    new_path = data.get('data_folder', '')
    
    if not new_path:
        return jsonify({'success': False, 'message': 'No path provided'}), 400
    
    # Expand ~ and other path shortcuts
    expanded_path = os.path.expanduser(new_path)
    
    # Validate path
    try:
        # Create directory if it doesn't exist
        os.makedirs(expanded_path, exist_ok=True)
        
        # Test if we can write to it
        test_file = os.path.join(expanded_path, '.test')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        
        # Save new config
        config = load_config()
        config['data_folder'] = expanded_path
        save_config(config)
        
        return jsonify({
            'success': True,
            'data_folder': expanded_path,
            'message': 'Data storage location updated successfully!'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Cannot access path: {str(e)}'
        }), 400

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """Get all transactions"""
    return jsonify(load_data())

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    """Add a new transaction"""
    transaction = request.json
    transactions = load_data()
    transaction['id'] = int(datetime.now().timestamp() * 1000)
    transactions.append(transaction)
    save_data(transactions)
    return jsonify(transaction), 201

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
def update_transaction(transaction_id):
    """Update a transaction"""
    updated_data = request.json
    transactions = load_data()
    
    for i, t in enumerate(transactions):
        if t['id'] == transaction_id:
            # Keep the ID, update everything else
            updated_data['id'] = transaction_id
            transactions[i] = updated_data
            save_data(transactions)
            return jsonify(updated_data), 200
    
    return jsonify({'success': False, 'message': 'Transaction not found'}), 404

@app.route('/api/transactions/bulk-update', methods=['POST'])
def bulk_update_transactions():
    """Bulk update multiple transactions"""
    data = request.json
    transaction_ids = data.get('transaction_ids', [])
    field = data.get('field')
    value = data.get('value')
    
    if not transaction_ids or not field or value is None:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
    
    # Validate field
    allowed_fields = ['category', 'type', 'status']
    if field not in allowed_fields:
        return jsonify({'success': False, 'message': 'Invalid field'}), 400
    
    transactions = load_data()
    updated_count = 0
    
    for t in transactions:
        if t['id'] in transaction_ids:
            t[field] = value
            updated_count += 1
    
    save_data(transactions)
    return jsonify({'success': True, 'updated': updated_count})

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
def delete_transaction(transaction_id):
    """Delete a transaction"""
    transactions = load_data()
    transactions = [t for t in transactions if t['id'] != transaction_id]
    save_data(transactions)
    return jsonify({'success': True})

@app.route('/api/budgets', methods=['GET'])
def get_budgets():
    """Get all budgets"""
    return jsonify(load_budgets())

@app.route('/api/budgets', methods=['POST'])
def set_budget():
    """Set a budget"""
    data = request.json
    budgets = load_budgets()
    budgets[data['category']] = data['amount']
    save_budgets(budgets)
    return jsonify(budgets)

@app.route('/api/budgets/<category>', methods=['DELETE'])
def delete_budget(category):
    """Delete a budget"""
    budgets = load_budgets()
    if category in budgets:
        del budgets[category]
        save_budgets(budgets)
    return jsonify({'success': True})

@app.route('/api/categories', methods=['GET'])
def get_categories():
    """Get all categories"""
    return jsonify(load_categories())

@app.route('/api/categories', methods=['POST'])
def add_category():
    """Add a new category"""
    data = request.json
    categories = load_categories()
    new_category = data.get('category', '').strip()
    
    if new_category and new_category not in categories:
        categories.append(new_category)
        save_categories(categories)
        return jsonify({'success': True, 'categories': categories})
    return jsonify({'success': False, 'message': 'Category already exists or invalid'}), 400

@app.route('/api/categories/<category>', methods=['DELETE'])
def delete_category(category):
    """Delete a category"""
    categories = load_categories()
    if category in categories:
        categories.remove(category)
        save_categories(categories)
        return jsonify({'success': True, 'categories': categories})
    return jsonify({'success': False, 'message': 'Category not found'}), 404

@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    """Import transactions from Excel file"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        transactions = load_data()
        imported_count = 0
        
        # Skip header row, start from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
                
            description = row[0] if row[0] else "No description"
            month = row[1] if row[1] else "January"
            amount = row[2] if row[2] else 0
            category = row[3] if row[3] else "Other"
            status = row[4] if row[4] else "Done"
            
            # Convert month to date (first day of month)
            try:
                if isinstance(month, str):
                    date_obj = datetime.strptime(f"{month} 2026", "%B %Y")
                    date_str = date_obj.strftime("%Y-%m-%d")
                else:
                    date_str = month.strftime("%Y-%m-%d") if hasattr(month, 'strftime') else "2026-01-01"
            except:
                date_str = "2026-01-01"
            
            # Determine type based on category and amount
            if "PayCheck" in str(category) or "Income" in str(category):
                trans_type = "income"
            else:
                trans_type = "expense"
            
            # Handle negative amounts
            try:
                amount_float = float(str(amount).replace('$', '').replace(',', '').replace('(', '-').replace(')', ''))
                amount_float = abs(amount_float)  # Store as positive, type determines income/expense
            except:
                amount_float = 0
            
            transaction = {
                'id': int(datetime.now().timestamp() * 1000) + imported_count,
                'type': trans_type,
                'category': category,
                'amount': amount_float,
                'date': date_str,
                'description': description,
                'status': status
            }
            
            transactions.append(transaction)
            imported_count += 1
        
        save_data(transactions)
        return jsonify({'success': True, 'imported': imported_count})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export', methods=['GET'])
def export_csv():
    """Export transactions as CSV"""
    transactions = load_data()
    
    # Helper to determine if type is income
    def is_income_type(t_type):
        return t_type in ['PayCheck', 'AdditionalIncome']
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Description', 'Month', 'Amount', 'Category', 'Status', 'Type'])
    
    for t in transactions:
        month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
        amount_display = t['amount'] if is_income_type(t['type']) else -t['amount']
        
        writer.writerow([
            t['date'],
            t.get('description', ''),
            month_name,
            amount_display,
            t['category'],
            t.get('status', 'Done'),
            t['type']
        ])
    
    output.seek(0)
    filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.csv"
    
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={filename}'
    }

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export transactions as Excel file"""
    transactions = load_data()
    
    # Helper to determine if type is income
    def is_income_type(t_type):
        return t_type in ['PayCheck', 'AdditionalIncome']
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Tracker"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Date', 'Description', 'Month', 'Amount', 'Category', 'Type', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    for row_idx, t in enumerate(transactions, 2):
        month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
        amount = t['amount'] if is_income_type(t['type']) else -t['amount']
        
        ws.cell(row=row_idx, column=1, value=t.get('id', ''))
        ws.cell(row=row_idx, column=2, value=t['date'])
        ws.cell(row=row_idx, column=3, value=t.get('description', ''))
        ws.cell(row=row_idx, column=4, value=month_name)
        ws.cell(row=row_idx, column=5, value=amount).number_format = '$#,##0.00'
        ws.cell(row=row_idx, column=6, value=t['category'])
        ws.cell(row=row_idx, column=7, value=t['type'])
        ws.cell(row=row_idx, column=8, value=t.get('status', 'Done'))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 12  # Date
    ws.column_dimensions['C'].width = 40  # Description
    ws.column_dimensions['D'].width = 12  # Month
    ws.column_dimensions['E'].width = 12  # Amount
    ws.column_dimensions['F'].width = 20  # Category
    ws.column_dimensions['G'].width = 10  # Type
    ws.column_dimensions['H'].width = 12  # Status
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    import socket
    
    import os


    port = int(os.environ.get('PORT', 5000))
    # Try ports 5000-5010 to find an available one
    # port = 5000
    # for p in range(5000, 5011):
    #     try:
    #         sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    #         sock.bind(('localhost', p))
    #         sock.close()
    #         port = p
    #         break
    #     except OSError:
    #         continue
    
    paths = get_data_paths()
    print(f"\n{'='*60}")
    print(f"💰 Expense Tracker Starting...")
    print(f"{'='*60}")
    print(f"📁 Data Storage Location: {paths['folder']}")
    print(f"⚙️  Config File: {CONFIG_FILE}")
    print(f"🌐 Access the app at: http://localhost:{port}")
    print(f"{'='*60}")
    print(f"\n⚠️  If port {port} doesn't work, port 5000 may be used by AirPlay.")
    print(f"    Solution: Use http://localhost:{port} or disable AirPlay Receiver")
    print(f"    in System Settings → General → AirDrop & Handoff\n")
    print(f"Press Ctrl+C to stop the server\n")

    app.run(host='0.0.0.0', port=port, debug=False)
    
    # app.run(debug=True, port=port, host='127.0.0.1')

from flask import Flask, render_template, request, jsonify, send_file
import json
import os
from datetime import datetime
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

# Import database helper
try:
    import db_helper
    DB_AVAILABLE = True
    print("✓ db_helper module imported successfully")
except Exception as e:
    print(f"✗ Failed to import db_helper: {e}")
    DB_AVAILABLE = False

app = Flask(__name__)

# Default categories (used for initialization only)
DEFAULT_CATEGORIES = [
    "Income", "Additional Income", "Housing", "AutoLoan", "Education", 
    "Healthcare", "Travel", "India Transfer", "House Expenses", 
    "HouseMaintenance", "Other"
]

def load_data():
    """Load transactions from PostgreSQL database"""
    try:
        if not DB_AVAILABLE:
            print("ERROR: Database helper not available")
            return []
        
        transactions = db_helper.load_transactions()
        print(f"✓ Loaded {len(transactions)} transactions from database")
        return transactions
        
    except Exception as e:
        print(f"ERROR loading transactions: {e}")
        traceback.print_exc()
        
        # Try to initialize database if tables don't exist
        try:
            print("Attempting to initialize database...")
            from init_db import init_database
            init_database()
            print("✓ Database initialized, trying to load again...")
            return db_helper.load_transactions()
        except Exception as init_error:
            print(f"ERROR initializing database: {init_error}")
            traceback.print_exc()
            return []

def load_categories():
    """Load categories from PostgreSQL database"""
    try:
        if not DB_AVAILABLE:
            print("ERROR: Database helper not available")
            return DEFAULT_CATEGORIES
        
        categories = db_helper.load_categories()
        
        if not categories:
            print("No categories found, initializing with defaults...")
            for cat in DEFAULT_CATEGORIES:
                db_helper.add_category(cat)
            return DEFAULT_CATEGORIES
        
        print(f"✓ Loaded {len(categories)} categories from database")
        return categories
        
    except Exception as e:
        print(f"ERROR loading categories: {e}")
        traceback.print_exc()
        return DEFAULT_CATEGORIES

@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """Get all transactions"""
    try:
        transactions = load_data()
        return jsonify(transactions)
    except Exception as e:
        error_msg = f"Error loading transactions: {str(e)}"
        print(f"ERROR in get_transactions: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 500

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    """Add a new transaction"""
    try:
        new_transaction = request.json
        
        # Validate required fields
        required_fields = ['date', 'description', 'amount', 'category', 'type']
        for field in required_fields:
            if field not in new_transaction:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Generate ID if not provided
        if 'id' not in new_transaction:
            import time
            new_transaction['id'] = int(time.time() * 1000)
        
        # Add status if not provided
        if 'status' not in new_transaction:
            new_transaction['status'] = 'Done'
        
        # Add to database
        db_helper.add_transaction(new_transaction)
        print(f"✓ Added transaction: {new_transaction['description']}")
        
        return jsonify(new_transaction), 201
        
    except Exception as e:
        error_msg = f"Error adding transaction: {str(e)}"
        print(f"ERROR in add_transaction: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
def update_transaction(transaction_id):
    """Update an existing transaction"""
    try:
        updated_transaction = request.json
        
        # Validate required fields
        required_fields = ['date', 'description', 'amount', 'category', 'type']
        for field in required_fields:
            if field not in updated_transaction:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        db_helper.update_transaction(transaction_id, updated_transaction)
        print(f"✓ Updated transaction {transaction_id}")
        
        updated_transaction['id'] = transaction_id
        return jsonify(updated_transaction)
        
    except Exception as e:
        error_msg = f"Error updating transaction: {str(e)}"
        print(f"ERROR in update_transaction: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
def delete_transaction(transaction_id):
    """Delete a transaction"""
    try:
        db_helper.delete_transaction(transaction_id)
        print(f"✓ Deleted transaction {transaction_id}")
        return jsonify({'message': 'Transaction deleted successfully'})
        
    except Exception as e:
        error_msg = f"Error deleting transaction: {str(e)}"
        print(f"ERROR in delete_transaction: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/transactions/bulk-update', methods=['POST'])
def bulk_update():
    """Bulk update transactions"""
    try:
        data = request.json
        transaction_ids = data.get('transaction_ids', [])
        field = data.get('field')
        value = data.get('value')
        
        if not transaction_ids or not field or not value:
            return jsonify({'error': 'Missing required fields'}), 400
        
        db_helper.bulk_update_transactions(transaction_ids, field, value)
        print(f"✓ Bulk updated {len(transaction_ids)} transactions")
        
        return jsonify({
            'message': f'Updated {len(transaction_ids)} transactions',
            'updated_count': len(transaction_ids)
        })
        
    except Exception as e:
        error_msg = f"Error in bulk update: {str(e)}"
        print(f"ERROR in bulk_update: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/transactions/bulk-delete', methods=['POST'])
def bulk_delete():
    """Bulk delete transactions"""
    try:
        data = request.json
        transaction_ids = data.get('transaction_ids', [])
        
        if not transaction_ids:
            return jsonify({'error': 'No transactions selected'}), 400
        
        for tid in transaction_ids:
            db_helper.delete_transaction(tid)
        
        print(f"✓ Bulk deleted {len(transaction_ids)} transactions")
        
        return jsonify({
            'message': f'Deleted {len(transaction_ids)} transactions',
            'deleted_count': len(transaction_ids)
        })
        
    except Exception as e:
        error_msg = f"Error in bulk delete: {str(e)}"
        print(f"ERROR in bulk_delete: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/categories', methods=['GET'])
def get_categories():
    """Get all categories"""
    try:
        categories = load_categories()
        return jsonify(categories)
    except Exception as e:
        error_msg = f"Error loading categories: {str(e)}"
        print(f"ERROR in get_categories: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 500

@app.route('/api/categories', methods=['POST'])
def add_category():
    """Add a new category"""
    try:
        data = request.json
        category_name = data.get('name', '').strip()
        
        if not category_name:
            return jsonify({'error': 'Category name required'}), 400
        
        success = db_helper.add_category(category_name)
        
        if success:
            print(f"✓ Added category: {category_name}")
            return jsonify({'message': 'Category added successfully', 'name': category_name}), 201
        else:
            return jsonify({'error': 'Category already exists or error occurred'}), 400
            
    except Exception as e:
        error_msg = f"Error adding category: {str(e)}"
        print(f"ERROR in add_category: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 400

@app.route('/api/export', methods=['GET'])
def export_csv():
    """Export transactions as CSV"""
    try:
        transactions = load_data()
        
        # Helper to determine if type is income
        def is_income_type(t_type):
            return t_type in ['PayCheck', 'AdditionalIncome']
        
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Description', 'Month', 'Amount', 'Category', 'Status', 'Type'])
        
        for t in transactions:
            month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
            amount_display = t['amount'] if is_income_type(t['type']) else -t['amount']
            
            writer.writerow([
                t['date'],
                t.get('description', ''),
                month_name,
                amount_display,
                t['category'],
                t.get('status', 'Done'),
                t['type']
            ])
        
        output.seek(0)
        filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        print(f"✓ Exported {len(transactions)} transactions to CSV")
        
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename={filename}'
        }
        
    except Exception as e:
        error_msg = f"Error exporting CSV: {str(e)}"
        print(f"ERROR in export_csv: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export transactions as Excel file"""
    try:
        transactions = load_data()
        
        # Helper to determine if type is income
        def is_income_type(t_type):
            return t_type in ['PayCheck', 'AdditionalIncome']
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Tracker"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['ID', 'Date', 'Description', 'Month', 'Amount', 'Category', 'Type', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for row_idx, t in enumerate(transactions, 2):
            month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
            amount = t['amount'] if is_income_type(t['type']) else -t['amount']
            
            ws.cell(row=row_idx, column=1, value=t.get('id', ''))
            ws.cell(row=row_idx, column=2, value=t['date'])
            ws.cell(row=row_idx, column=3, value=t.get('description', ''))
            ws.cell(row=row_idx, column=4, value=month_name)
            ws.cell(row=row_idx, column=5, value=amount).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=6, value=t['category'])
            ws.cell(row=row_idx, column=7, value=t['type'])
            ws.cell(row=row_idx, column=8, value=t.get('status', 'Done'))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        print(f"✓ Exported {len(transactions)} transactions to Excel")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        error_msg = f"Error exporting Excel: {str(e)}"
        print(f"ERROR in export_excel: {error_msg}")
        traceback.print_exc()
        return jsonify({'error': error_msg}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint for debugging"""
    try:
        # Test database connection
        transactions = db_helper.load_transactions()
        categories = db_helper.load_categories()
        
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'transactions_count': len(transactions),
            'categories_count': len(categories),
            'db_helper_available': DB_AVAILABLE
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'db_helper_available': DB_AVAILABLE
        }), 500

if __name__ == '__main__':
    import os
    
    # Print startup diagnostics
    print("\n" + "="*60)
    print("EXPENSE TRACKER - STARTUP DIAGNOSTICS")
    print("="*60)
    
    # Check database connection
    try:
        print("\n1. Testing database connection...")
        test_transactions = load_data()
        print(f"   ✓ Database connected - {len(test_transactions)} transactions found")
        
        test_categories = load_categories()
        print(f"   ✓ Categories loaded - {len(test_categories)} categories found")
        
    except Exception as e:
        print(f"   ✗ Database connection failed: {e}")
        print("\n   Please check:")
        print("   - PostgreSQL is running")
        print("   - Database 'expense_tracker' exists")
        print("   - Password in db_helper.py is correct")
        print("   - Run: python init_db.py")
    
    print("\n" + "="*60)
    
    port = int(os.environ.get('PORT', 5001))
    print(f"\n🚀 Starting Flask app on http://localhost:{port}")
    print("="*60 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=False)

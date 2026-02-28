from flask import Flask, render_template, request, jsonify, send_file
import json
import os
from datetime import datetime
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import database helper
import db_helper

app = Flask(__name__)

# Default categories (used for initialization only)
DEFAULT_CATEGORIES = [
    "Income", "Additional Income", "Housing", "AutoLoan", "Education", 
    "Healthcare", "Travel", "India Transfer", "House Expenses", 
    "HouseMaintenance", "Other"
]

def load_data():
    """Load transactions from PostgreSQL database"""
    try:
        return db_helper.load_transactions()
    except Exception as e:
        print(f"Error loading transactions: {e}")
        # Initialize database if it doesn't exist
        try:
            from init_db import init_database
            init_database()
            return []
        except:
            return []

def save_data(transactions):
    """
    Note: With PostgreSQL, we don't save all transactions at once.
    Individual transactions are saved via add/update/delete operations.
    This function is kept for compatibility but does nothing.
    """
    pass

def load_categories():
    """Load categories from PostgreSQL database"""
    try:
        categories = db_helper.load_categories()
        if not categories:
            # Initialize with default categories
            for cat in DEFAULT_CATEGORIES:
                db_helper.add_category(cat)
            return DEFAULT_CATEGORIES
        return categories
    except Exception as e:
        print(f"Error loading categories: {e}")
        return DEFAULT_CATEGORIES

@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """Get all transactions"""
    transactions = load_data()
    return jsonify(transactions)

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    """Add a new transaction"""
    try:
        new_transaction = request.json
        
        # Generate ID if not provided
        if 'id' not in new_transaction:
            import time
            new_transaction['id'] = int(time.time() * 1000)
        
        # Add to database
        db_helper.add_transaction(new_transaction)
        
        return jsonify(new_transaction), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
def update_transaction(transaction_id):
    """Update an existing transaction"""
    try:
        updated_transaction = request.json
        db_helper.update_transaction(transaction_id, updated_transaction)
        
        updated_transaction['id'] = transaction_id
        return jsonify(updated_transaction)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
def delete_transaction(transaction_id):
    """Delete a transaction"""
    try:
        db_helper.delete_transaction(transaction_id)
        return jsonify({'message': 'Transaction deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/transactions/bulk-update', methods=['POST'])
def bulk_update():
    """Bulk update transactions"""
    try:
        data = request.json
        transaction_ids = data.get('transaction_ids', [])
        field = data.get('field')
        value = data.get('value')
        
        if not transaction_ids or not field or not value:
            return jsonify({'error': 'Missing required fields'}), 400
        
        db_helper.bulk_update_transactions(transaction_ids, field, value)
        
        return jsonify({
            'message': f'Updated {len(transaction_ids)} transactions',
            'updated_count': len(transaction_ids)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/transactions/bulk-delete', methods=['POST'])
def bulk_delete():
    """Bulk delete transactions"""
    try:
        data = request.json
        transaction_ids = data.get('transaction_ids', [])
        
        if not transaction_ids:
            return jsonify({'error': 'No transactions selected'}), 400
        
        for tid in transaction_ids:
            db_helper.delete_transaction(tid)
        
        return jsonify({
            'message': f'Deleted {len(transaction_ids)} transactions',
            'deleted_count': len(transaction_ids)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/categories', methods=['GET'])
def get_categories():
    """Get all categories"""
    categories = load_categories()
    return jsonify(categories)

@app.route('/api/categories', methods=['POST'])
def add_category():
    """Add a new category"""
    try:
        data = request.json
        category_name = data.get('name', '').strip()
        
        if not category_name:
            return jsonify({'error': 'Category name required'}), 400
        
        success = db_helper.add_category(category_name)
        
        if success:
            return jsonify({'message': 'Category added successfully', 'name': category_name}), 201
        else:
            return jsonify({'error': 'Category already exists or error occurred'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/export', methods=['GET'])
def export_csv():
    """Export transactions as CSV"""
    transactions = load_data()
    
    # Helper to determine if type is income
    def is_income_type(t_type):
        return t_type in ['PayCheck', 'AdditionalIncome']
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Description', 'Month', 'Amount', 'Category', 'Status', 'Type'])
    
    for t in transactions:
        month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
        amount_display = t['amount'] if is_income_type(t['type']) else -t['amount']
        
        writer.writerow([
            t['date'],
            t.get('description', ''),
            month_name,
            amount_display,
            t['category'],
            t.get('status', 'Done'),
            t['type']
        ])
    
    output.seek(0)
    filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.csv"
    
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={filename}'
    }

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export transactions as Excel file"""
    transactions = load_data()
    
    # Helper to determine if type is income
    def is_income_type(t_type):
        return t_type in ['PayCheck', 'AdditionalIncome']
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Tracker"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Date', 'Description', 'Month', 'Amount', 'Category', 'Type', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    for row_idx, t in enumerate(transactions, 2):
        month_name = datetime.strptime(t['date'], '%Y-%m-%d').strftime('%B')
        amount = t['amount'] if is_income_type(t['type']) else -t['amount']
        
        ws.cell(row=row_idx, column=1, value=t.get('id', ''))
        ws.cell(row=row_idx, column=2, value=t['date'])
        ws.cell(row=row_idx, column=3, value=t.get('description', ''))
        ws.cell(row=row_idx, column=4, value=month_name)
        ws.cell(row=row_idx, column=5, value=amount).number_format = '$#,##0.00'
        ws.cell(row=row_idx, column=6, value=t['category'])
        ws.cell(row=row_idx, column=7, value=t['type'])
        ws.cell(row=row_idx, column=8, value=t.get('status', 'Done'))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 12  # Date
    ws.column_dimensions['C'].width = 40  # Description
    ws.column_dimensions['D'].width = 12  # Month
    ws.column_dimensions['E'].width = 12  # Amount
    ws.column_dimensions['F'].width = 20  # Category
    ws.column_dimensions['G'].width = 10  # Type
    ws.column_dimensions['H'].width = 12  # Status
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"expense-tracker-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
